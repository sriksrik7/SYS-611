################################ Start of file ##################################
# File Name: Game.py
# Description: This is file contains all methods to simulate men singles wimbledon
#              tennis tournament
#################################################################################

from openpyxl import Workbook, load_workbook
import random
from Player import Tennis_Player

################################ Start of function ##############################
# Function Name: is_player1_win_coin_flip
# Description: Function to determine if player 1 has won the toss or not
# Return: returns true or false
#################################################################################
def is_player1_win_coin_flip():
    heads = 1
    coin = random.randint(0,1)
    if coin == 1:
        return True
    else:
        return False
################################ End of function ################################


################################ Start of function ##############################
# Function Name: process_input_statistics_data
# Description: function to read all players statistics data for a given worksheet
#              title name
# Return: None
#################################################################################
def process_input_statistics_data(worksheet_title_name):
    # Load the work book:
    wb_obj = load_workbook('Statistics_Data.xlsx')
    # Read the user provided workbook
    stat_sheet = wb_obj[worksheet_title_name]

    cellnum = 1
    for player in players:
        cellnum += 1
        player.set_name(stat_sheet[f'A{cellnum}'].value)
        player.set_stat(player.ACE, stat_sheet[f'D{cellnum}'].value * 100)
        player.set_stat(player.DOUBLE_FAULT, stat_sheet[f'E{cellnum}'].value * 100)
        player.set_stat(player.FIRST_SERVE, stat_sheet[f'F{cellnum}'].value * 100)
        player.set_stat(player.FIRST_SERVE_WON, stat_sheet[f'G{cellnum}'].value * 100)
        player.set_stat(player.SECOND_SERVE_WON, stat_sheet[f'H{cellnum}'].value * 100)
        player.set_stat(player.BREAK_POINT_SAVED, stat_sheet[f'I{cellnum}'].value * 100)
        player.set_stat(player.SERVICE_POINTS_WON, stat_sheet[f'J{cellnum}'].value * 100)
        player.set_stat(player.SERVICE_GAMES_WON, stat_sheet[f'K{cellnum}'].value * 100)
        player.set_stat(player.ACE_AGAINST, stat_sheet[f'L{cellnum}'].value * 100)
        player.set_stat(player.FIRST_SERVE_RET_WON, stat_sheet[f'M{cellnum}'].value * 100)
        player.set_stat(player.SECOND_SERVE_RET_WON, stat_sheet[f'N{cellnum}'].value * 100)
        player.set_stat(player.BREAK_POINTS_WON, stat_sheet[f'O{cellnum}'].value * 100)
        player.set_stat(player.RET_POINTS_WON, stat_sheet[f'P{cellnum}'].value * 100)
        player.set_stat(player.RET_GAMES_WON, stat_sheet[f'Q{cellnum}'].value * 100)

    # Close the workbook after reading
    wb_obj.close()

################################ End of function ################################


################################ Start of function ##############################
# Function Name: get_player_index
# Description: function to get player index from a list of 32 players
# Return: returns player index from a list of 32 players
#################################################################################
def get_player_index(p_name):
    ret_player_index = int()
    # Loop through all 32 players to get the index for a given player name
    for p_idx in range(32):
        if (players[p_idx].get_name() == p_name):
            ret_player_index = p_idx
            break
    return ret_player_index
################################ End of function ################################


################################ Start of function ##############################
# Function Name: simulate_set
# Description: Function to simulate a set in Tennis
# Return: returns a string contains the wining player's rank/index, player 1 score,
#         player 2 score, is any player injured, injured player index.
#         Eg: Wining player rank/index, player 1 score , player 2 score,
#             is any player injured, injured player index.
#         Eg: 2,4,6,False,None
#         Eg: 2,3,4,True,23
#################################################################################
def simulate_set(player1_idx, player2_idx):

    # Fetch player 1 stats weightage
    player1_weightage = players[player1_idx].weightage_calculation()
    # Fetch player 2 stats weightage
    player2_weightage = players[player2_idx].weightage_calculation()

    # holds player1 and player2's game points for a set
    player1_gamepoints = 0
    player2_gamepoints = 0

    # holds player1 and player2's injury information
    is_player1_injured = False
    is_player2_injured = False
    is_any_player_injured = False
    player_injured_idx = None

    # Loop through to simulate all 12 sets
    for game_index in range(12):
        rand_fp = random.uniform(0.0, 100.0)
        # Even serves are serviced by player 1
        if (game_index % 2) == 0:
            if rand_fp <= player1_weightage:
                player1_gamepoints += 1
            else:
                player2_gamepoints += 1
        # ODD serves are serviced by player 2
        else:
            if rand_fp <= player2_weightage:
                player2_gamepoints += 1
            else:
                player1_gamepoints += 1

        # Check if player 1 is injured during the sets
        injury_rand_p1 = random.randint(0, 10000)
        if injury_rand_p1 <= 17:
            is_player1_injured = True
        else:
            # Check if player 2 is injured during the sets
            injury_rand_p2 = random.randint(0, 10000)
            if injury_rand_p2 <= 17:
                is_player2_injured = True

        # If player 1 is injured forfeit the match for player 1, and make player 2 as a winner
        if is_player1_injured == True:
            player_won_idx = player2_idx
            player_injured_idx = player1_idx
            is_any_player_injured = True
            break

        # If player 2 is injured forfeit the match for player 2, and make player 1 as a winner
        if is_player2_injured == True:
            player_won_idx = player1_idx
            player_injured_idx = player2_idx
            is_any_player_injured = True
            break

        # If a player has scored 6 points and if he is 1 set ahead of the opponent,
        # then the set is complete. Hence break from the for loop to finalise the set.
        # Eg: 6-0,6-1,6-2,6-3 and 6-4
        if player1_gamepoints == 6 and game_index >= 5 and game_index != 10:
            break
        # Eg: 0-6,1-6,2-6,3-6 and 4-6
        elif player2_gamepoints == 6 and game_index >= 5 and game_index != 10:
            break

    # If none of the players are injured, process player win determination
    if is_any_player_injured == False:
        # Check if both players scored 6-6
        # Based on higher weightage of a player, determine the tie breaker
        if player1_gamepoints == player2_gamepoints:
            if player1_weightage > player2_weightage:
                player_won_idx = player1_idx
                # Eg: 7-6
                player1_gamepoints += 1
            else:
                player_won_idx = player2_idx
                # Eg: 6-7
                player2_gamepoints += 1
        elif player1_gamepoints > player2_gamepoints:
            player_won_idx = player1_idx
        else:
            # if player2_gamepoints > player1_gamepoints
            player_won_idx = player2_idx

    # String contains the winner rank/index followed by the set result
    # Eg: Wining player rank/index, player 1 score , player 2 score, is any player injured, injured player index.
    # Eg: 2,4,6,False,None
    # Eg: 2,3,4,True,23
    ret_string = f"{player_won_idx},{player1_gamepoints},{player2_gamepoints},{is_any_player_injured},{player_injured_idx}"
    # print(ret_string)
    return ret_string
################################ End of function ################################


################################ Start of function ##############################
# Function Name: write_results_to_cell
# Description: This function to writes a tennis set results and determines if a
#              match is forfeited
# Return: returns a string containing the match forfeit information along with
#         wining player's index
#################################################################################
def write_results_to_cell( str_result, wr_cell_num, is_player_index_flipped, ply1_index, ply2_index, set_result_ws ):

    is_match_forfeited =  False
    player_won_index = None
    # Split the string the read set details
    # Eg: 2,4,6,False,None
    # Eg: 2,3,4,True,23
    part_str_set_result = str_result.split(',')

    # Cell number to write results for player 2
    next_cell = f"{wr_cell_num[0]}{int(wr_cell_num[1:])+1}"

    # If player 1 has lost the toss, flip and write the results in assigned cell numbers
    # of player 1 and player 2
    if is_player_index_flipped == True:
        # Write player2 result in current cell, since the player1 become player2 due to coin toss
        set_result_ws[wr_cell_num] = int(part_str_set_result[2])
        # Write player1 result in next cell, since the player2 become player1 due to coin toss
        set_result_ws[next_cell] = int(part_str_set_result[1])
        # If the match is forfeited, process the string to determine the winner and
        # write the results in respective cell number
        if part_str_set_result[3] == "True":
            is_match_forfeited = True
            # if player1 is injured, write x to next cell
            if part_str_set_result[4] == ply1_index:
                set_result_ws[next_cell] = "X"
                # Set player2 has winner, since player 1 is injured
                player_won_index = ply2_index
            # if player2 is injured, write x to current cell
            else:
                set_result_ws[wr_cell_num] = "X"
                # Set player1 has winner, since player 2 is injured
                player_won_index = ply1_index
    else:
        # Don't flip write the results as it is designated cell numbers.
        set_result_ws[wr_cell_num] = int(part_str_set_result[1])
        set_result_ws[next_cell] = int(part_str_set_result[2])

        # If the match is forfeited, process the string to determine the winner and
        # write the results in respective cell number
        if part_str_set_result[3] == "True":
            is_match_forfeited = True
            # if player1 is injured, write x to current cell
            if part_str_set_result[4] == ply1_index:
                set_result_ws[wr_cell_num] = "X"
                # Set player2 has winner, since player 1 is injured
                player_won_index = ply2_index
            # if player2 is injured, write x to next cell
            else:
                set_result_ws[next_cell] = "X"
                # Set player1 has winner, since player 2 is injured
                player_won_index = ply1_index

    # Format the return string to include both match forfeit and winning player's
    # index inforamtion
    return_result_str = f"{is_match_forfeited},{player_won_index}"
    return return_result_str
################################ End of function ################################


################################ Start of function ##############################
# Function Name: simulate_matches
# Description: Function to simulate tennis matches
# Return: returns a list of winners
#################################################################################
def simulate_matches(cell_num, num_matches, match_detail_list, mtc_result_ws):

    # List to hold the winners for given stage of simulated matches
    list_of_winners = []
    cell_aphabet = ["D", "E", "F", "G", "H"]

    mtc_result_ws[f'D{cell_num}'] = "Set 1"
    mtc_result_ws[f'E{cell_num}'] = "Set 2"
    mtc_result_ws[f'F{cell_num}'] = "Set 3"
    mtc_result_ws[f'G{cell_num}'] = "Set 4"
    mtc_result_ws[f'H{cell_num}'] = "Set 5"
    mtc_result_ws[f'I{cell_num}'] = "Player Won"

    for match_list_idx in range(num_matches):
        # Increment the cell number
        cell_num += 1

        #Is player1 and player2 flipped at coin toss
        #if flipped player 2 is servicing first else player 1 is servicing first
        is_player_idx_flipped = False

        # Default the flag "is_match_abandoned" to False
        is_match_abandoned = False

        # Fetch player 1 and player 2's indices/ranks from the match schedule list
        if is_player1_win_coin_flip():
            is_player_idx_flipped = False
            p1_index = match_detail_list[match_list_idx]['player1_index']
            p2_index = match_detail_list[match_list_idx]['player2_index']
        else:
            is_player_idx_flipped = True
            p2_index = match_detail_list[match_list_idx]['player1_index']
            p1_index = match_detail_list[match_list_idx]['player2_index']

        # Holds player1 and player 2's win count
        p1_win_count = 0
        p2_win_count = 0
        # Simulate first 3 sets to check if any player has already won all three sets.
        for match_played in range(3):
            set_result = simulate_set(p1_index,p2_index)
            part_set_result = set_result.split(',')
            # print(f"p1_index:{p1_index} = {part_set_result[0]}")
            if p1_index == int(part_set_result[0]):
                p1_win_count += 1
            else:
                p2_win_count += 1

            # Write the set results to the cells
            is_match_abnd_details = write_results_to_cell(set_result, f'{cell_aphabet[match_played]}{cell_num}', is_player_idx_flipped, p1_index, p2_index, mtc_result_ws)
            is_match_abnd_info = is_match_abnd_details.split(',')

            # If match is abandoned, break the simulation for rest of the match
            if is_match_abnd_info[0] == "True":
                is_match_abandoned = True
                break

        # If match is not abandoned, check if any player has already won 3 sets.
        if is_match_abandoned == False:
            # Check if player 1 has already won 3 sets.
            if p1_win_count == 3:
                mtc_result_ws[f'I{cell_num}'] = f"{players[p1_index].get_name()}"
                list_of_winners.append(players[p1_index].get_name())
                # Increment the cell number for next match
                cell_num += 1
                continue
            # Check if player 2 has already won 3 sets.
            elif p2_win_count == 3:
                mtc_result_ws[f'I{cell_num}'] = f"{players[p2_index].get_name()}"
                list_of_winners.append(players[p2_index].get_name())
                # Increment the cell number for next match
                cell_num += 1
                continue
        # If match is abandoned, write the player win inforamtion in designated cell
        # and add the player name to list winners for this stage of matches
        else:
            mtc_result_ws[f'I{cell_num}'] = f"{players[int(is_match_abnd_info[1])].get_name()}"
            list_of_winners.append(players[int(is_match_abnd_info[1])].get_name())
            # Increment the cell number for next match
            cell_num += 1
            continue


        # If none of the player has won 3 sets continue to simulate 4th set
        set4_result = simulate_set(p1_index,p2_index)
        part_set4_result = set4_result.split(',')

        if p1_index == int(part_set4_result[0]):
            p1_win_count += 1
        else:
            p2_win_count += 1

        # Write the set results to the cells
        is_match_abnd_details = write_results_to_cell(set4_result, f'G{cell_num}', is_player_idx_flipped, p1_index, p2_index, mtc_result_ws)
        is_match_abnd_info = is_match_abnd_details.split(',')
        # print(is_match_abnd_info)

        # If match is abandoned, set is_match_abandoned flag to True
        if is_match_abnd_info[0] == "True":
            is_match_abandoned = True

        # If match is not abandoned, check if any player has already won 3 sets out of 4 sets
        if is_match_abandoned == False:
            # Check if player 1 has already won 3 sets.
            if p1_win_count == 3:
                mtc_result_ws[f'I{cell_num}'] = f"{players[p1_index].get_name()}"
                list_of_winners.append(players[p1_index].get_name())
                # Increment the cell number for next match
                cell_num += 1
                continue
            # Check if player 2 has already won 3 sets.
            elif p2_win_count == 3:
                mtc_result_ws[f'I{cell_num }'] = f"{players[p2_index].get_name()}"
                list_of_winners.append(players[p2_index].get_name())
                # Increment the cell number for next match
                cell_num += 1
                continue
        # If match is abandoned, write the player win inforamtion in designated cell
        # and add the player name to list winners for this stage of matches
        else:
            mtc_result_ws[f'I{cell_num}'] = f"{players[int(is_match_abnd_info[1])].get_name()}"
            list_of_winners.append(players[int(is_match_abnd_info[1])].get_name())
            # Increment the cell number for next match and continue to simulate next match
            cell_num += 1
            continue

        # If none of the player has won 3 sets continue to simulate final set
        set5_result = simulate_set(p1_index,p2_index)
        part_set5_result = set5_result.split(',')
        # print(is_match_abnd_info)

        if p1_index == int(part_set5_result[0]):
            p1_win_count += 1
        else:
            p2_win_count += 1

        # Write the set results to the cells
        is_match_abnd_details = write_results_to_cell(set5_result, f'H{cell_num}', is_player_idx_flipped, p1_index, p2_index, mtc_result_ws)
        is_match_abnd_info = is_match_abnd_details.split(',')

        # If match is abandoned, set is_match_abandoned flag to True
        if is_match_abnd_info[0] == "True":
            is_match_abandoned = True

        # If match is not abandoned, determine who won the match and write the winning
        # player's name in the designated cell, also add the player's name to list winners
        # for this stage of matches
        if is_match_abandoned == False:
            # Check if player 1 has already won 3 sets.
            if p1_win_count == 3:
                mtc_result_ws[f'I{cell_num}'] = f"{players[p1_index].get_name()}"
                list_of_winners.append(players[p1_index].get_name())
            # Check if player 2 has already won 3 sets.
            elif p2_win_count == 3:
                mtc_result_ws[f'I{cell_num}'] = f"{players[p2_index].get_name()}"
                list_of_winners.append(players[p2_index].get_name())
        # If match is abandoned, write the player win information in designated cell
        # and add the player name to list winners for this stage of matches
        else:
            mtc_result_ws[f'I{cell_num}'] = f"{players[int(is_match_abnd_info[1])].get_name()}"
            list_of_winners.append(players[int(is_match_abnd_info[1])].get_name())

        # Increment the cell number for next match
        cell_num += 1

    return list_of_winners
################################ End of function ################################


################################ Start of function ##############################
# Function Name: simulate_tournament
# Description: function to simulate all matches for men singles tennis tournament
# Return: returns the  final winner name
#################################################################################
# Simulate Men's singles Wimbledon tournament
def simulate_tournament(simulation_num, workbook_obj, sim_year):

    # Variables to hold winners for each stage of tournament
    first_round_winners = []
    scnd_round_winners = []
    quarter_final_winners = []
    semi_final_winners = []

    # Cell start index numbers for each stage of tournament
    rd1_cell_index = 2
    rd2_cell_index = 37
    qf_cell_index = 56
    sf_cell_index = 67
    final_cell_index = 74
    champ_cell_index = 78

    #########################################################
    ####### Populate and process first round matches #######

    ####### Bracket seeding for 32 players #######
    # Top 16 out of 32 players are played againts the lower 16 players.
    # Lower 16 players are randomly allocated to play against the top 16 players
    # This is to avoid strong players playing against each other in the intial stages of tournament.
    # This type seeding is followed in Wembledom tournament to make the final games challening.

    sim_num = simulation_num +1
    sch_res_ws = workbook_obj.create_sheet("Mysheet")
    sch_res_ws.title = f"Sim{sim_num}_results"

    sch_res_ws[f'A{rd1_cell_index-1}'] = "Wimbledon Round 1"
    sch_res_ws[f'A{rd1_cell_index}'] = "Match"
    sch_res_ws[f'B{rd1_cell_index}'] = "Rank"
    sch_res_ws[f'C{rd1_cell_index}'] = "Player Name"

    # Create a match details list to incude match number and player 1 ranking.
    # This match details list can be used to subscript each player statistics data based on their ranking.
    # Holds match number along with player1 and player 2's index/rank details
    first_rd_match_list = []
    rd1_cell = rd1_cell_index
    for index in range(16):
        rd1_cell+=1
        sch_res_ws[f'A{rd1_cell}'] = index+1
        sch_res_ws[f'B{rd1_cell}'] = index+1
        sch_res_ws[f'C{rd1_cell}'] = players[index].get_name()
        first_rd_match_list.append({'match_index':(index+1),'player1_index':(index)})
        rd1_cell+=1

    # Randomly assign the lower 16 players to play against the top 16 players.
    processed_numbers = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31]
    cell_sch = rd1_cell_index
    for sch_index in range(16):
        rank = random.choice(processed_numbers)
        processed_numbers.remove(rank)
        cell_sch+=2
        sch_res_ws[f'B{cell_sch}'] = rank+1
        sch_res_ws[f'C{cell_sch}'] = players[rank].get_name()
        first_rd_match_list[sch_index]['player2_index'] = rank

    # print(first_rd_match_list)
    # Save the bracket seeding schedule into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    # Simulate matches for first round
    first_round_winners = simulate_matches(rd1_cell_index, 16, first_rd_match_list, sch_res_ws)

    # Save the results of first round matches
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    #########################################################
    ####### Populate and process second round matches #######

    sch_res_ws[f'A{rd2_cell_index-1}'] = "Wimbledon Round 2"
    sch_res_ws[f'A{rd2_cell_index}'] = "Match"
    sch_res_ws[f'B{rd2_cell_index}'] = "Rank"
    sch_res_ws[f'C{rd2_cell_index}'] = "Player Name"

    # Schedule second round matches
    scnd_rd_match_list = []
    rd2_top_cell = rd2_cell_index
    for win_rd1_top_index in range(8):
        rd2_top_cell+=1
        # Fetch player indices from the list of 32 players to simulate second round matches
        pl1_index = get_player_index(first_round_winners[win_rd1_top_index])

        # Populate second round schedule from first round winners.
        sch_res_ws[f'A{rd2_top_cell}'] = win_rd1_top_index+1
        sch_res_ws[f'B{rd2_top_cell}'] = pl1_index+1
        sch_res_ws[f'C{rd2_top_cell}'] = first_round_winners[win_rd1_top_index]
        # Generate a list of dictionaries for second round matches
        scnd_rd_match_list.append({'match_index':(win_rd1_top_index+1),'player1_index':(pl1_index)})
        rd2_top_cell+=1

    rd2_low_cell = (rd2_cell_index+17)
    for win_rd1_low_index in reversed(range(8)):
        rd2_low_cell-=1
        pl2_index = get_player_index(first_round_winners[win_rd1_low_index+8])
        sch_res_ws[f'B{rd2_low_cell}'] = pl2_index+1
        sch_res_ws[f'C{rd2_low_cell}'] = first_round_winners[win_rd1_low_index+8]
        # Generate a list of dictionaries for second round matches
        scnd_rd_match_list[win_rd1_low_index]['player2_index'] = pl2_index
        rd2_low_cell-=1

    # Save the second round schedule into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    # Simulate matches for second round
    scnd_round_winners = simulate_matches(rd2_cell_index, 8, scnd_rd_match_list, sch_res_ws)

    # Save the second round results into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    ##########################################################
    ####### Populate and process quarter final matches #######
    sch_res_ws[f'A{qf_cell_index-1}'] = "Wimbledon Quarter Finals"
    sch_res_ws[f'A{qf_cell_index}'] = "Match"
    sch_res_ws[f'B{qf_cell_index}'] = "Rank"
    sch_res_ws[f'C{qf_cell_index}'] = "Player Name"

    # Schedule quarter final matches
    qf_match_list = []
    qf_cell = qf_cell_index
    win_rd2_index = 0
    for index in range(4):
        qf_cell+=1
        # Fetch player indices from the list of 32 players to simulate quarter final matches
        pl1_index = get_player_index(scnd_round_winners[win_rd2_index])
        pl2_index = get_player_index(scnd_round_winners[win_rd2_index+1])

        # Populate quarter final schedule from second round winners.
        sch_res_ws[f'A{qf_cell}'] = index+1
        sch_res_ws[f'B{qf_cell}'] = pl1_index+1
        sch_res_ws[f'C{qf_cell}'] = scnd_round_winners[win_rd2_index]
        sch_res_ws[f'B{qf_cell+1}'] = pl2_index+1
        sch_res_ws[f'C{qf_cell+1}'] = scnd_round_winners[win_rd2_index+1]
        # Generate a list of dictionaries for quarter final matches
        qf_match_list.append({'match_index':(index+1),'player1_index':(pl1_index),'player2_index':(pl2_index)})
        qf_cell+=1
        win_rd2_index+=2

    # Save the quarter final schedule into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    # Simulate matches for quarter final
    quarter_final_winners = simulate_matches(qf_cell_index, 4, qf_match_list, sch_res_ws)

    # Save the quarter final results into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    ##########################################################
    ####### Populate and process semi final matches #######
    sch_res_ws[f'A{sf_cell_index-1}'] = "Wimbledon Semi Final"
    sch_res_ws[f'A{sf_cell_index}'] = "Match"
    sch_res_ws[f'B{sf_cell_index}'] = "Rank"
    sch_res_ws[f'C{sf_cell_index}'] = "Player Name"

    # Schedule semi final matches
    sf_match_list = []
    sf_cell = sf_cell_index
    win_qf_index = 0
    for index in range(2):
        sf_cell+=1
        # Fetch player indices from the list of 32 players to simulate semi final matches
        pl1_index = get_player_index(quarter_final_winners[win_qf_index])
        pl2_index = get_player_index(quarter_final_winners[win_qf_index+1])

        # Populated semi final schedule from quarter final winners.
        sch_res_ws[f'A{sf_cell}'] = index+1
        sch_res_ws[f'B{sf_cell}'] = pl1_index+1
        sch_res_ws[f'C{sf_cell}'] = quarter_final_winners[win_qf_index]
        sch_res_ws[f'B{sf_cell+1}'] = pl2_index+1
        sch_res_ws[f'C{sf_cell+1}'] = quarter_final_winners[win_qf_index+1]
        # Generate a list of dictionaries for semi final matches
        sf_match_list.append({'match_index':(index+1),'player1_index':(pl1_index),'player2_index':(pl2_index)})
        sf_cell+=1
        win_qf_index+=2

    # Save the semi final schedule into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    # Simulate matches for semi final
    semi_final_winners = simulate_matches(sf_cell_index, 2, sf_match_list, sch_res_ws)

    # Save the semi final results into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    ##########################################################
    ####### Populate and process final match #################
    sch_res_ws[f'A{final_cell_index-1}'] = "Wimbledon Final"
    sch_res_ws[f'A{final_cell_index}'] = "Match"
    sch_res_ws[f'B{final_cell_index}'] = "Rank"
    sch_res_ws[f'C{final_cell_index}'] = "Player Name"

    # Schedule final matche
    final_match_list = []

    # Fetch player indices from the list of 32 players to simulate final match
    pl1_index = get_player_index(semi_final_winners[0])
    pl2_index = get_player_index(semi_final_winners[1])

    # Populate final schedule from semi final winners.
    sch_res_ws[f'A{final_cell_index+1}'] = 1
    sch_res_ws[f'B{final_cell_index+1}'] = pl1_index+1
    sch_res_ws[f'C{final_cell_index+1}'] = semi_final_winners[0]
    sch_res_ws[f'B{final_cell_index+2}'] = pl2_index+1
    sch_res_ws[f'C{final_cell_index+2}'] = semi_final_winners[1]
    # Generate a list of dictionaries for final matche
    final_match_list.append({'match_index':(1),'player1_index':(pl1_index),'player2_index':(pl2_index)})

    # Save the final schedule into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    # Simulate matches for final
    final_winner = simulate_matches(final_cell_index, 1, final_match_list, sch_res_ws)

    # Record the simulate wimbledon champion
    # Fetch player indices from the list of 32 players to simulate final match
    Champ_pl_index = get_player_index(final_winner[0])
    sch_res_ws[f'A{champ_cell_index}'] = "Wimbledon Champion"
    sch_res_ws[f'B{champ_cell_index}'] = "Rank"
    sch_res_ws[f'C{champ_cell_index}'] = "Player Name"

    sch_res_ws[f'B{champ_cell_index+1}'] = Champ_pl_index+1
    sch_res_ws[f'C{champ_cell_index+1}'] = final_winner[0]

    # print(f"The wimbledon Champion is {final_winner[0]}")
    # Save the final results into a spread sheet.
    workbook_obj.save(f"Wimbledon_Model_Results_{sim_year}.xlsx")

    return final_winner[0]

################################ End of function ################################


#################################################################################
########################     Main  Starts Here       ############################
#################################################################################

# Global Variables
players = []
players = [Tennis_Player() for i in range(32)]

# Take input from the user
print("Enter the year to simulate Wimbledon tennis tournament")
sheet_input = input("Choose from following years: 2019, 2018, 2017:\n")

if sheet_input == "2019":
    sheet_title = "2019"
elif sheet_input == "2018":
    sheet_title = "2018"
elif sheet_input == "2017":
    sheet_title = "2017"
else:
    raise ValueError(f"Invalid input {sheet_input}, Please try again...")

# Take input from the user to perform number of simulations
total_simulations = int(input("Enter the number of simulations to be performed:\nNote: Simulations are limited to 30\n -->"))

if total_simulations > 30:
    raise ValueError(f"The provided number of simualtions is greater than 30, Please try again...")

print("Please be patient it takes about 3 minutes to simulate")

# Process input statistics data for a given sheet name from file Statistics_Data.xlsx
process_input_statistics_data(sheet_title)

# Create a new work book to record all results of each simulated tournament
sch_res_wb = Workbook()

tournament_winners_dict = {}
# simulate multiple number of tournaments for user provided input simulations
for sim_idx in range(total_simulations):
    tournament_winner = simulate_tournament(sim_idx, sch_res_wb, sheet_title)

    if tournament_winner in tournament_winners_dict:
        tournament_winners_dict[tournament_winner] += 1
    else:
        tournament_winners_dict.update({tournament_winner: 1})

# Re-use the work "Sheet" for winnerSimInfo
tournament_winners_ws = sch_res_wb["Sheet"]
tournament_winners_ws.title = "WinnersSimInfo"
tournament_winners_ws[f'A1'] = "Tournament Winner Name"
tournament_winners_ws[f'B1'] = "Number of times won"

winner_idx = 1
for winner in tournament_winners_dict:
    winner_idx +=1
    tournament_winners_ws[f'A{winner_idx}'] = winner
    tournament_winners_ws[f'B{winner_idx}'] = tournament_winners_dict[winner]

# Save the spread sheet after updating the WinnersInfo.
sch_res_wb.save(f"Wimbledon_Model_Results_{sheet_title}.xlsx")

print(tournament_winners_dict)

# Close the workbook after writing simulated data
sch_res_wb.close()
#################################################################################
########################         End of Main         ############################
#################################################################################

################################## End of file ##################################